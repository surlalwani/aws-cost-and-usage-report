import email
from botocore.exceptions import ClientError
from botocore.exceptions import ProfileNotFound
from email.mime.multipart import MIMEMultipart       
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

import argparse
import boto3
import csv
from collections import defaultdict
from datetime import datetime
import datetime
from openpyxl import Workbook
import glob
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Color,Alignment

def lambda_handler(event, context):
    # TODO implement
    
    parser = argparse.ArgumentParser()
    parser.add_argument('--days', type=int, default=30)
    args = parser.parse_args()
    
    
    now = datetime.datetime.utcnow()
    start = (now - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    end = now.strftime('%Y-%m-%d')

    access_key_id=""
    secret_access_key=""
    ec2 = boto3.client('ec2','us-east-1',aws_access_key_id=access_key_id, aws_secret_access_key=secret_access_key)
    
    response = ec2.describe_instances (Filters=[
     {
      'Name': 'instance-state-name',
      'Values': ['running']
     }
    ])
    for r in response['Reservations']:
    ### Creating a CSV file and inputting all EC2 isntances information in it.
        cnt=1
        csv_ob=open("/tmp/demo.csv","w",newline='')
        csv_w=csv.writer(csv_ob)
        csv_w.writerow(['Name',"Instance_Id",'Instance_Type','PrivateIpAddress','PublicIP','Time(hrs)'])
        response=ec2.describe_instances (Filters=[
            {
                'Name': 'instance-state-name',
                'Values': ['running']
            }
        ])['Reservations']
    name_list = []
    for each_item in response:
        for instances in each_item['Instances']:
            diff_time = datetime.datetime.now(tz=datetime.timezone.utc) - instances['LaunchTime'] 
            diff_time = diff_time.total_seconds()/(3600.0)
            for tags in instances['Tags']:
                if tags['Key'] == 'Name':
                    name_list.append(tags['Value'])
                    csv_w.writerow([tags['Value'],instances['InstanceId'],instances['InstanceType'],instances['PrivateIpAddress'],instances['PublicIpAddress'],diff_time])
                    cnt+=1

    csv_ob.close()

    ###  Converting above CSV file into excel sheet AWSresourcesReport.xlsx
    
    wb = Workbook()
    ws = wb.active
    with open('/tmp/demo.csv', 'r') as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save('/tmp/AWSresourcesReport.xlsx')
    mr = ws.max_row
    mc = ws.max_column
     ### Formatting rows and columns according to our requirement 
    for cell in ws["mr:mc"]:
        cell.font = Font(size=11)

    for cell in ws["1:1"]:
        cell.font = Font(size=12)
        cell.style = "Accent1"
        cell.alignment = Alignment(wrapText="True", horizontal="center")

##Creating asheet name EC2 instance in Excel file
    ss_sheet = wb['Sheet']
    ss_sheet.title = 'EC2Instances'
    wb.save('/tmp/AWSresourcesReport.xlsx')
    df = pd.read_excel("/tmp/AWSresourcesReport.xlsx",index_col=0)   
    
  ## Creating a bar chart based on Time Consumption and name of EC2 instance.
    shift_pro = df.loc[:,"Time(hrs)"]
    print(shift_pro)
    shift_pro.plot(kind='bar')
    
    plt.xlabel("Name of Instances")
    plt.ylabel("Time(hr)")
    plt.show(block=False)
    shift_pro.plot(kind='bar')
    fig = plt.gcf()
    fig.set_size_inches(5.5, 7.5,forward=True)
  ## Bar chart is created and we save it as a png file then uploaded it to excel file
    fig.savefig('/tmp/fig1.png',dpi=100)\
  ## Loading workbook and uploading bar chart 
    wk = load_workbook('/tmp/AWSresourcesReport.xlsx')
    wm = wb.worksheets[0]
    print(wm)
    wm.merge_cells('L7:L9')
    img = openpyxl.drawing.image.Image('/tmp/fig1.png')
    cell = wm.cell(row=5, column=5)
    chart = BarChart3D()
    wm.add_image(img,'j2')
    
    wb.save('/tmp/AWSresourcesReport.xlsx')

    #~~~~~~S3 list~~~~~~~~~~~~~~~~~~~
    regions = ['us-east-1']
    measurable_metrics = [
        ('BucketSizeBytes', 'StandardStorage'),
        ('NumberOfObjects', 'AllStorageTypes'),
    ]
    
    s3 = boto3.resource('s3','us-east-1',aws_access_key_id=access_key_id, aws_secret_access_key=secret_access_key)
    cnt=1
    csv_ob=open("/tmp/s3bucketlist.csv","w",newline='')
    csv_w=csv.writer(csv_ob)
    csv_w.writerow(["S_NO","Name",'Bucket_size(GB)'])
    for bucket in s3.buckets.all():
     my_bucket = s3.Bucket(bucket.name)
##Converting Bytes into MB (Size of Bucket)
     bytes = sum([object.size for object in my_bucket.objects.all()])
     print(f'total bucket size: {bytes//1000/1024/1024}')
     csv_w.writerow([cnt,bucket.name,f'{bytes//1000/1024/1024}'])
     cnt+=1
    csv_ob.close()
## Creating a 2nd sheet in Excel report named S3 Bucket
    wb.create_sheet('S3 Buckets',2)
    mr = ws.max_row
    mc = ws.max_column
## Formatting columns and rows font size 
    for cell in ws["mr:mc"]:
        cell.font = Font(size=11)
    for cell in ws["1:1"]:
        cell.font = Font(size=12)
        cell.style = "Accent1"
        cell.alignment = Alignment(wrapText="True", horizontal="center")
    
    wb.save('/tmp/AWSresourcesReport.xlsx')
    print(wb.sheetnames)
    with open('/tmp/s3bucketlist.csv', 'r') as f:
        for row in csv.reader(f):
            wb['S3 Buckets'].append(row)
    wb.save('/tmp/AWSresourcesReport.xlsx')

    # ~~~~~~~~~~~~~~~~~~~~~AMI ~~~~~~~~~~~~~~~
    client = boto3.client('ec2','us-east-1',aws_access_key_id=access_key_id, aws_secret_access_key=secret_access_key)
    response = client.describe_images(Owners=['self'])
    cnt=1
    csv_ob=open("/tmp/AMI.csv","w",newline='')
    csv_w=csv.writer(csv_ob)
    csv_w.writerow(["S_NO","AMI_Id",'Count'])
    
    for ami in response['Images']:
        print (ami['ImageId'])
        csv_w.writerow([cnt,ami['ImageId']])
        cnt+=1
    csv_ob.close()
    total=cnt-1
    wh = load_workbook('/tmp/AWSresourcesReport.xlsx')
## Creating a 3rd sheet in Excel report named AMI count
    wb.create_sheet('AMI Count',3)
## Formaatting font size of rows and columns 
    for cell in ws["mr:mc"]:
        cell.font = Font(size=11)
    for cell in ws["1:1"]:
        cell.font = Font(size=12)
        cell.style = "Accent1"
        cell.alignment = Alignment(wrapText="True", horizontal="center")
    wb.save('/tmp/AWSresourcesReport.xlsx')
    print(wb.sheetnames)
    
    with open('/tmp/AMI.csv', 'r') as f:
        for row in csv.reader(f):
            wb['AMI Count'].append(row)
            wb['AMI Count']['C2'] = total
    wb.save('/tmp/AWSresourcesReport.xlsx')

    #~~~~~~~~~~~~~~~~~~~Snapshot~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    Snaps=ec2.describe_snapshots(OwnerIds=['self'])
    print(Snaps['Snapshots'])
    cnt=1
    csv_ob=open("/tmp/snapshot.csv","w",newline='')
    csv_w=csv.writer(csv_ob)
    csv_w.writerow(['Snapshot Id','Size'])
    
    for snapshot in Snaps['Snapshots']:
        csv_w.writerow([snapshot['SnapshotId'],snapshot['VolumeSize']])
        cnt+=1
    csv_ob.close()
    wh = load_workbook('/tmp/AWSresourcesReport.xlsx')
## Creating a 4th sheet in Excel report named Snapshot
    wb.create_sheet('Snapshots',4)
## Formaatting font size of rows and columns
    for cell in ws["mr:mc"]:
        cell.font = Font(size=11)
    for cell in ws["1:1"]:
        cell.font = Font(size=12)
        cell.style = "Accent1"
        cell.alignment = Alignment(wrapText="True", horizontal="center")
    wb.save('/tmp/AWSresourcesReport.xlsx')
    print(wb.sheetnames)
    
    with open('/tmp/snapshot.csv', 'r') as f:
        for row in csv.reader(f):
            wb['Snapshots'].append(row)
    wb.save('/tmp/AWSresourcesReport.xlsx')


     ## Uploading excel report to S3 bucket
    s3_client = boto3.client('s3',aws_access_key_id=access_key_id, aws_secret_access_key=secret_access_key)
    response = s3_client.upload_file('/tmp/AWSresourcesReport.xlsx', 'ishan-cicd-images', 'AWSresourcesReport.xlsx')
 ## Sending report using simple email service 
    SENDER = ""
    RECIPIENT = ""
    AWS_REGION = "us-east-1"
    SUBJECT = "AWS Resources Report"
    BUCKET_NAME = '' # replace with your bucket name
    KEY = 'AWSresourcesReport.xlsx' # replace with your object key
    # current_time = datetime.now()
    time = now.strftime('%m-%d-%Y')
    key =  'AWS Resources Report' + str(time) + '.xlsx'
    print(now.strftime('%m/%d/%Y'))
    
    
    s3 = boto3.resource('s3',region_name=AWS_REGION,aws_access_key_id=access_key_id, aws_secret_access_key=secret_access_key)
    
# Downloading file from S3 Bucket 
    s3.Bucket(BUCKET_NAME).download_file(KEY, '/tmp/'+ key )
## Using Simple email service for sending mails.
    client = boto3.client('ses',region_name=AWS_REGION,aws_access_key_id=access_key_id, aws_secret_access_key=secret_access_key)
    msg = MIMEMultipart()
    # Add subject, from and to lines.
    msg['Subject'] = SUBJECT 
    msg['From'] = SENDER 
    msg['To'] = RECIPIENT
    BODY_TEXT="Hello,Please find the AWS Resources Report attached for the date of {}".format(time)

    textpart = MIMEText(BODY_TEXT)
    msg.attach(textpart)
    
    att = MIMEApplication(open('/tmp/{}'.format(key), 'rb').read())
    # filen = "AWS Resources Report" + time + ".xlsx"
    att.add_header('Content-Disposition','attachment',filename=key)
    msg.attach(att)
    print(msg)
    try:
        response = client.send_raw_email(
            Source=msg['From'],
            Destinations=msg['To'].split(","),
            RawMessage={ 'Data':msg.as_string() }
        )
    except ClientError as e:
        print(e.response['Error']['Message'])
    else:
        print("Email sent! Message ID:",response['MessageId'])
        
    
    return {
        'statusCode': 200,
        'body': 'Hello from Lambda!'
    }


